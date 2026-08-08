import io
import openpyxl
import pytest
from app.models.session import Session
from app.models.student import Student
from app.models.problem import Problem
from app.models.code_execution import CodeExecution
from app.services.excel_report_service import ExcelReportService
from app.services.report_service import ReportService

def setup_test_session(client, mode="practice"):
    """Helper to set up a test session and student."""
    s_res = client.post("/api/v1/teacher/session/create", json={
        "teacher_name": "Dr. Excel", "teacher_email": "excel@edu.com", "college": "Tech",
        "department": "CS", "subject": "Data Lab", "title": "Excel Session",
        "language": "python", "mode": mode
    })
    s_data = s_res.get_json()["data"]
    t_token = s_data["teacher_token"]
    s_id = s_data["session"]["id"]
    s_code = s_data["session"]["session_code"]

    j_res = client.post("/api/v1/student/session/join", json={
        "session_code": s_code, "name": "Excel Student", "roll_number": "XL_001",
        "department": "CS", "year": "2nd", "section": "B"
    })
    j_data = j_res.get_json()["data"]
    st_token = j_data["student_token"]
    st_id = j_data["student"]["id"]

    return t_token, st_token, s_id, st_id, s_code

# ------------------------------------------------------------------------------
# EXCEL REPORTING TEST SUITE
# ------------------------------------------------------------------------------

def test_excel_report_generation(client, db):
    """1. Test ExcelReportService.generate_excel_report(session_id)."""
    _, _, s_id, _, _ = setup_test_session(client)
    excel_bytes, filename = ExcelReportService.generate_excel_report(s_id)
    assert excel_bytes is not None
    assert len(excel_bytes) > 0
    assert filename.endswith(".xlsx")

def test_excel_workbook_validity(client, db):
    """2. Test openpyxl can load generated Excel workbook bytes successfully."""
    _, _, s_id, _, _ = setup_test_session(client)
    excel_bytes, _ = ExcelReportService.generate_excel_report(s_id)
    wb = openpyxl.load_workbook(io.BytesIO(excel_bytes))
    assert wb is not None

def test_excel_all_7_sheets_exist(client, db):
    """3. Test all 7 required sheets exist in Excel workbook."""
    _, _, s_id, _, _ = setup_test_session(client)
    excel_bytes, _ = ExcelReportService.generate_excel_report(s_id)
    wb = openpyxl.load_workbook(io.BytesIO(excel_bytes))

    expected_sheets = ["Summary", "Students", "Code Executions", "Activity", "AI Analysis", "Problem", "Errors"]
    for sheet in expected_sheets:
        assert sheet in wb.sheetnames

def test_excel_summary_sheet(client, db):
    """4. Test Summary sheet contents."""
    _, _, s_id, _, _ = setup_test_session(client)
    excel_bytes, _ = ExcelReportService.generate_excel_report(s_id)
    wb = openpyxl.load_workbook(io.BytesIO(excel_bytes))
    ws = wb["Summary"]
    assert "CodeSphere AI" in str(ws.cell(row=1, column=1).value)

def test_excel_students_sheet(client, db):
    """5. Test Students sheet headers and data."""
    _, _, s_id, _, _ = setup_test_session(client)
    excel_bytes, _ = ExcelReportService.generate_excel_report(s_id)
    wb = openpyxl.load_workbook(io.BytesIO(excel_bytes))
    ws = wb["Students"]
    headers = [cell.value for cell in ws[1]]
    assert "Name" in headers
    assert "Roll Number" in headers
    assert "Progress" in headers

def test_excel_code_executions_sheet(client, db):
    """6. Test Code Executions sheet headers."""
    _, _, s_id, _, _ = setup_test_session(client)
    excel_bytes, _ = ExcelReportService.generate_excel_report(s_id)
    wb = openpyxl.load_workbook(io.BytesIO(excel_bytes))
    ws = wb["Code Executions"]
    headers = [cell.value for cell in ws[1]]
    assert "Student" in headers
    assert "Status" in headers
    assert "Exit Code" in headers

def test_excel_download_api(client, db):
    """7. Test GET /api/v1/teacher/session/{id}/report/excel API endpoint."""
    t_token, _, s_id, _, s_code = setup_test_session(client)
    res = client.get(f"/api/v1/teacher/session/{s_id}/report/excel", headers={"X-Teacher-Token": t_token})
    assert res.status_code == 200
    assert res.content_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    disposition = res.headers.get("Content-Disposition", "")
    assert f"codesphere_session_{s_code}_report.xlsx" in disposition

def test_report_status_api_both_pdf_and_excel(client, db):
    """8. Test GET /api/v1/teacher/session/{id}/report/status returns status for both PDF and Excel."""
    t_token, _, s_id, _, _ = setup_test_session(client)
    res = client.get(f"/api/v1/teacher/session/{s_id}/report/status", headers={"X-Teacher-Token": t_token})
    assert res.status_code == 200
    data = res.get_json()["data"]
    assert "pdf" in data
    assert "excel" in data

def test_report_summary_api(client, db):
    """9. Test GET /api/v1/teacher/session/{id}/report/summary API contract."""
    t_token, _, s_id, _, _ = setup_test_session(client)
    res = client.get(f"/api/v1/teacher/session/{s_id}/report/summary", headers={"X-Teacher-Token": t_token})
    assert res.status_code == 200
    data = res.get_json()["data"]
    assert "downloads" in data
    assert "pdf" in data["downloads"]
    assert "excel" in data["downloads"]

def test_sensitive_credentials_never_exposed_in_excel(client, db):
    """10. Test that sensitive API keys, tokens, and credentials NEVER appear in Excel workbook."""
    _, _, s_id, _, _ = setup_test_session(client)
    excel_bytes, _ = ExcelReportService.generate_excel_report(s_id)
    wb = openpyxl.load_workbook(io.BytesIO(excel_bytes))

    sensitive_patterns = ["OPENAI_API_KEY", "ONLINE_COMPILER_API_KEY", "JWT_SECRET", "postgres://"]
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows(values_only=True):
            for cell_val in row:
                if cell_val:
                    cell_str = str(cell_val)
                    for pattern in sensitive_patterns:
                        assert pattern not in cell_str
