import io
import json
from datetime import datetime, timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.models.session import Session
from app.models.student import Student
from app.models.problem import Problem
from app.models.code_snapshot import CodeSnapshot
from app.models.code_execution import CodeExecution
from app.models.activity_event import ActivityEvent
from app.models.ai_review import AIReview
from app.services.redis_service import get_redis_client
from app.logger import api_logger, error_logger

class ExcelReportService:
    """Excel Report Generator Service using openpyxl."""

    @classmethod
    def generate_excel_report(cls, session_id: int) -> tuple[bytes, str]:
        """Generate complete 7-sheet Excel workbook report."""
        session = Session.query.get(session_id)
        if not session:
            raise ValueError(f"Session ID {session_id} not found")

        students = Student.query.filter_by(session_id=session_id).order_by(Student.joined_at.asc()).all()
        problem = Problem.query.filter_by(session_id=session_id).order_by(Problem.created_at.desc()).first()
        executions = CodeExecution.query.filter_by(session_id=session_id).order_by(CodeExecution.created_at.asc()).all()
        activities = ActivityEvent.query.filter_by(session_id=session_id).order_by(ActivityEvent.created_at.asc()).all()
        ai_reviews = AIReview.query.filter_by(session_id=session_id).order_by(AIReview.created_at.asc()).all()

        total_students = len(students)
        total_runs = len(executions)
        successful_runs = sum(1 for e in executions if e.exit_code == 0)
        failed_runs = total_runs - successful_runs

        wb = Workbook()
        
        # Styles
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        title_font = Font(name="Calibri", size=16, bold=True, color="1E293B")
        sub_font = Font(name="Calibri", size=11, bold=True, color="2563EB")

        def style_header(ws, row=1):
            for cell in ws[row]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")

        def autofit(ws):
            for col in ws.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = get_column_letter(col[0].column)
                ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

        # ----------------------------------------------------------------------
        # SHEET 1: Summary
        # ----------------------------------------------------------------------
        ws1 = wb.active
        ws1.title = "Summary"

        ws1.append(["CodeSphere AI - Session Analytics Summary Report"])
        ws1.cell(row=1, column=1).font = title_font
        ws1.append([])

        ws1.append(["SESSION METADATA", ""])
        ws1.cell(row=3, column=1).font = sub_font
        summary_metadata = [
            ["Session ID", session.id],
            ["Session Code", session.session_code],
            ["Title", session.title],
            ["Teacher Name", session.teacher_name],
            ["Teacher Email", session.teacher_email],
            ["College", session.college],
            ["Department", session.department],
            ["Subject", session.subject],
            ["Mode", session.mode],
            ["Language", session.language],
            ["Created Time", session.created_at.strftime("%Y-%m-%d %H:%M:%S UTC") if session.created_at else "N/A"],
            ["End Time", session.ended_at.strftime("%Y-%m-%d %H:%M:%S UTC") if session.ended_at else "N/A"]
        ]
        for row in summary_metadata:
            ws1.append(row)

        ws1.append([])
        ws1.append(["CLASSROOM METRICS", ""])
        ws1.cell(row=17, column=1).font = sub_font
        metrics = [
            ["Total Students", total_students],
            ["Total Code Runs", total_runs],
            ["Successful Runs", successful_runs],
            ["Failed Runs", failed_runs],
            ["Total Activity Events", len(activities)],
            ["AI Reviews Generated", len(ai_reviews)]
        ]
        for row in metrics:
            ws1.append(row)
        autofit(ws1)

        # ----------------------------------------------------------------------
        # SHEET 2: Students
        # ----------------------------------------------------------------------
        ws2 = wb.create_sheet(title="Students")
        ws2.append(["Name", "Roll Number", "Department", "Year", "Section", "Joined At", "Last Active", "Time Spent", "Status", "Progress", "AI Score", "Code Quality", "Executions", "Successful Runs", "Failed Runs", "Stuck Status", "Warnings"])
        style_header(ws2)

        for s in students:
            s_reviews = [r for r in ai_reviews if r.student_id == s.id]
            prog = "N/A"
            ai_score = "N/A"
            quality = "N/A"
            for r in s_reviews:
                if r.progress is not None:
                    prog = f"{r.progress}%"
                if r.code_quality is not None:
                    quality = f"{r.code_quality}/100"
                    ai_score = quality

            s_runs = [e for e in executions if e.student_id == s.id]
            s_succ = sum(1 for e in s_runs if e.exit_code == 0)
            s_fail = len(s_runs) - s_succ

            joined_str = s.joined_at.strftime("%H:%M:%S") if s.joined_at else "N/A"
            active_str = s.last_active.strftime("%H:%M:%S") if s.last_active else "N/A"

            ws2.append([
                s.name, s.roll_number, s.department, s.year, s.section,
                joined_str, active_str, "N/A", s.status, prog, ai_score,
                quality, len(s_runs), s_succ, s_fail, "False", "None"
            ])
        autofit(ws2)

        # ----------------------------------------------------------------------
        # SHEET 3: Code Executions
        # ----------------------------------------------------------------------
        ws3 = wb.create_sheet(title="Code Executions")
        ws3.append(["Student", "Roll Number", "Language", "Status", "Exit Code", "Execution Time", "Memory", "Error", "Created At"])
        style_header(ws3)

        for e in executions:
            st = next((s for s in students if s.id == e.student_id), None)
            st_name = st.name if st else "Unknown"
            st_roll = st.roll_number if st else "N/A"
            c_time = e.created_at.strftime("%Y-%m-%d %H:%M:%S") if e.created_at else "N/A"

            ws3.append([
                st_name, st_roll, e.language, e.status, e.exit_code,
                e.execution_time or "0s", e.memory or "0MB", e.error or "", c_time
            ])
        autofit(ws3)

        # ----------------------------------------------------------------------
        # SHEET 4: Activity
        # ----------------------------------------------------------------------
        ws4 = wb.create_sheet(title="Activity")
        ws4.append(["Student", "Roll Number", "Event", "Metadata", "Timestamp"])
        style_header(ws4)

        for a in activities:
            st = next((s for s in students if s.id == a.student_id), None)
            st_name = st.name if st else "Unknown"
            st_roll = st.roll_number if st else "N/A"
            a_time = a.created_at.strftime("%Y-%m-%d %H:%M:%S") if a.created_at else "N/A"

            ws4.append([
                st_name, st_roll, a.event_type, a.event_metadata or "", a_time
            ])
        autofit(ws4)

        # ----------------------------------------------------------------------
        # SHEET 5: AI Analysis
        # ----------------------------------------------------------------------
        ws5 = wb.create_sheet(title="AI Analysis")
        ws5.append(["Student", "Roll Number", "Progress", "Confidence", "Code Quality", "Logic Score", "Readability", "Efficiency", "Current Stage", "Summary", "Suggestions", "Stuck Status"])
        style_header(ws5)

        for r in ai_reviews:
            st = next((s for s in students if s.id == r.student_id), None)
            st_name = st.name if st else "Unknown"
            st_roll = st.roll_number if st else "N/A"

            ws5.append([
                st_name, st_roll, r.progress or "N/A", r.confidence or "N/A",
                r.code_quality or "N/A", "N/A", "N/A", "N/A",
                r.analysis_type, r.summary or "", r.suggestions_json or "", "False"
            ])
        autofit(ws5)

        # ----------------------------------------------------------------------
        # SHEET 6: Problem
        # ----------------------------------------------------------------------
        ws6 = wb.create_sheet(title="Problem")
        ws6.append(["Attribute", "Details"])
        style_header(ws6)

        if problem:
            ws6.append(["Problem Title", problem.title])
            ws6.append(["Description", problem.description])
            ws6.append(["Constraints", problem.constraints or "None"])
            ws6.append(["Input Format", problem.input_format or "None"])
            ws6.append(["Output Format", problem.output_format or "None"])
            ws6.append(["Sample Input", problem.sample_input or "None"])
            ws6.append(["Sample Output", problem.sample_output or "None"])
            ws6.append(["Teacher Reference Solution", problem.reference_solution or "None"])
        else:
            ws6.append(["Mode", "Practice Mode (No structured problem statement assigned)"])
        autofit(ws6)

        # ----------------------------------------------------------------------
        # SHEET 7: Errors
        # ----------------------------------------------------------------------
        ws7 = wb.create_sheet(title="Errors")
        ws7.append(["Student", "Roll Number", "Error Type", "Error Message", "Count", "First Occurrence", "Last Occurrence", "AI Explanation"])
        style_header(ws7)

        err_execs = [e for e in executions if e.exit_code != 0 or e.error]
        for e in err_execs:
            st = next((s for s in students if s.id == e.student_id), None)
            st_name = st.name if st else "Unknown"
            st_roll = st.roll_number if st else "N/A"
            e_time = e.created_at.strftime("%Y-%m-%d %H:%M:%S") if e.created_at else "N/A"

            ws7.append([
                st_name, st_roll, e.status or "Error", e.error or "Non-zero exit code",
                1, e_time, e_time, "Syntax or logic mismatch"
            ])
        autofit(ws7)

        # Compile Excel Workbook to Bytes
        out_buf = io.BytesIO()
        wb.save(out_buf)
        excel_bytes = out_buf.getvalue()
        out_buf.close()

        filename = f"codesphere_session_{session.session_code}_report.xlsx"

        # Cache in Redis
        try:
            r = get_redis_client()
            r.set(f"session:{session_id}:excel_report", excel_bytes, ex=86400)
            r.set(f"session:{session_id}:excel_report_status", "ready", ex=86400)
            r.set(f"session:{session_id}:excel_filename", filename, ex=86400)
        except Exception as e:
            error_logger.warning(f"Failed to cache Excel report in Redis: {str(e)}")

        return excel_bytes, filename

    @classmethod
    def get_cached_excel_report(cls, session_id: int) -> tuple[bytes, str]:
        """Retrieve temporarily cached Excel report bytes from Redis."""
        try:
            r = get_redis_client()
            excel_bytes = r.get(f"session:{session_id}:excel_report")
            filename = r.get(f"session:{session_id}:excel_filename")
            if excel_bytes:
                fn_str = filename.decode("utf-8") if isinstance(filename, bytes) else (filename or f"codesphere_session_{session_id}_report.xlsx")
                return excel_bytes, fn_str
        except Exception:
            pass

        return cls.generate_excel_report(session_id)
