import os
import io
import json
import uuid
from datetime import datetime, timezone, timedelta
from sqlalchemy import func
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from app.extensions import db
from app.models.session import Session
from app.models.student import Student
from app.models.session_student import SessionStudent
from app.models.problem import Problem
from app.models.code_execution import CodeExecution
from app.models.activity_event import ActivityEvent
from app.models.ai_review import AIReview
from app.models.report_job import ReportJob
from app.services.report_service import ReportService
from app.services.excel_report_service import ExcelReportService
from app.logger import api_logger, error_logger

EXPORT_DIR = "./exports/reports"
os.makedirs(EXPORT_DIR, exist_ok=True)

class ReportV2Service:
    """Service for V2 Filtered Any-Time Report Generation."""

    @classmethod
    def create_report_job(cls, filter_type: str, filter_params: dict, teacher_id: int = None, report_format: str = "both") -> ReportJob:
        """Create a new report job record in PostgreSQL."""
        job_id = f"rep_{uuid.uuid4().hex[:12]}"
        job = ReportJob(
            job_id=job_id,
            teacher_id=teacher_id,
            session_id=filter_params.get("session_id"),
            student_id=filter_params.get("student_id"),
            filter_type=filter_type,
            filter_params_json=json.dumps(filter_params),
            report_format=report_format,
            status="pending"
        )
        job.save()
        return job

    @classmethod
    def execute_report_job(cls, job_id: str):
        """Execute PDF and Excel report generation for a job."""
        job = ReportJob.query.filter_by(job_id=job_id).first()
        if not job:
            return None

        try:
            job.status = "processing"
            db.session.commit()

            params = json.loads(job.filter_params_json or "{}")
            filter_type = job.filter_type

            # If specific session filter, delegate to existing high-fidelity session report generators
            if filter_type == "session" and job.session_id:
                pdf_bytes, pdf_name = ReportService.generate_session_report(job.session_id)
                excel_bytes, excel_name = ExcelReportService.generate_excel_report(job.session_id)

                pdf_path = os.path.join(EXPORT_DIR, f"{job_id}.pdf")
                excel_path = os.path.join(EXPORT_DIR, f"{job_id}.xlsx")

                with open(pdf_path, "wb") as f:
                    f.write(pdf_bytes)
                with open(excel_path, "wb") as f:
                    f.write(excel_bytes)

                job.file_path_pdf = pdf_path
                job.file_path_excel = excel_path
                job.status = "ready"
                job.completed_at = datetime.now(timezone.utc)
                db.session.commit()
                return job

            # Compute Date Range for Today, Monthly, Custom, or Student filters
            now_utc = datetime.now(timezone.utc)
            if filter_type == "today":
                start_date = datetime(now_utc.year, now_utc.month, now_utc.day, 0, 0, 0, tzinfo=timezone.utc)
                end_date = datetime(now_utc.year, now_utc.month, now_utc.day, 23, 59, 59, tzinfo=timezone.utc)
                title = f"Today Report ({start_date.strftime('%d %b %Y')})"
            elif filter_type == "monthly":
                month_str = params.get("month") # e.g. "2026-08"
                if month_str:
                    try:
                        dt = datetime.strptime(month_str, "%Y-%m")
                        year, month = dt.year, dt.month
                    except Exception:
                        year, month = now_utc.year, now_utc.month
                else:
                    year, month = now_utc.year, now_utc.month
                start_date = datetime(year, month, 1, 0, 0, 0, tzinfo=timezone.utc)
                # Next month first day minus 1 sec
                if month == 12:
                    next_m = datetime(year + 1, 1, 1, 0, 0, 0, tzinfo=timezone.utc)
                else:
                    next_m = datetime(year, month + 1, 1, 0, 0, 0, tzinfo=timezone.utc)
                end_date = next_m - timedelta(seconds=1)
                title = f"Monthly Report ({start_date.strftime('%B %Y')})"
            elif filter_type == "custom":
                s_str = params.get("start_date")
                e_str = params.get("end_date")
                try:
                    start_date = datetime.strptime(s_str, "%Y-%m-%d").replace(tzinfo=timezone.utc)
                except Exception:
                    start_date = now_utc - timedelta(days=7)
                try:
                    end_date = datetime.strptime(e_str, "%Y-%m-%d").replace(hour=23, minute=59, second=59, tzinfo=timezone.utc)
                except Exception:
                    end_date = now_utc
                title = f"Custom Report ({start_date.strftime('%d/%m/%Y')} - {end_date.strftime('%d/%m/%Y')})"
            elif filter_type == "student" and job.student_id:
                st = Student.query.get(job.student_id)
                title = f"Student History Report: {st.name} ({st.roll_number})" if st else "Student History Report"
                start_date = datetime(2020, 1, 1, tzinfo=timezone.utc)
                end_date = now_utc + timedelta(days=1)
            else:
                start_date = now_utc - timedelta(days=30)
                end_date = now_utc
                title = "Filtered Performance Report"

            # Query PostgreSQL historical sessions in date range
            sess_query = Session.query.filter(Session.created_at >= start_date, Session.created_at <= end_date)
            if job.teacher_id:
                sess_query = sess_query.filter(Session.teacher_id == job.teacher_id)
            sessions = sess_query.order_by(Session.created_at.desc()).all()
            session_ids = [s.id for s in sessions]

            # Query executions, participations, students
            if session_ids:
                executions = CodeExecution.query.filter(CodeExecution.session_id.in_(session_ids)).all()
                participations = SessionStudent.query.filter(SessionStudent.session_id.in_(session_ids)).all()
                ai_reviews = AIReview.query.filter(AIReview.session_id.in_(session_ids)).all()
            else:
                executions = []
                participations = []
                ai_reviews = []

            total_sessions = len(sessions)
            total_students = len(set(p.student_id for p in participations)) if participations else len(set(e.student_id for e in executions if e.student_id))
            total_runs = len(executions)
            successful_runs = sum(1 for e in executions if e.exit_code == 0)
            failed_runs = max(0, total_runs - successful_runs)

            avg_progress = int(sum(p.progress for p in participations) / len(participations)) if participations else 0
            avg_ai_score = round(sum(p.ai_score for p in participations) / len(participations), 1) if participations else 0.0
            avg_code_quality = round(sum(p.code_quality for p in participations) / len(participations), 1) if participations else 0.0

            pdf_path = os.path.join(EXPORT_DIR, f"{job_id}.pdf")
            excel_path = os.path.join(EXPORT_DIR, f"{job_id}.xlsx")

            # Generate PDF
            cls._build_pdf_report(pdf_path, title, total_sessions, total_students, total_runs, successful_runs, failed_runs, avg_progress, avg_ai_score, avg_code_quality, sessions, participations)
            # Generate Excel
            cls._build_excel_report(excel_path, title, total_sessions, total_students, total_runs, successful_runs, failed_runs, avg_progress, avg_ai_score, avg_code_quality, sessions, participations)

            job.file_path_pdf = pdf_path
            job.file_path_excel = excel_path
            job.status = "ready"
            job.completed_at = datetime.now(timezone.utc)
            db.session.commit()
            return job

        except Exception as e:
            error_logger.error(f"Report execution failed for job {job_id}: {str(e)}")
            job.status = "failed"
            job.error_message = str(e)
            db.session.commit()
            return job

    @classmethod
    def _build_pdf_report(cls, filepath, title, total_sessions, total_students, total_runs, successful_runs, failed_runs, avg_progress, avg_ai_score, avg_code_quality, sessions, participations):
        doc = SimpleDocTemplate(filepath, pagesize=letter, rightMargin=36, leftMargin=36, topMargin=36, bottomMargin=36)
        styles = getSampleStyleSheet()

        title_style = ParagraphStyle('DocTitle', parent=styles['Heading1'], fontSize=18, leading=22, textColor=colors.HexColor('#1E293B'), spaceAfter=12)
        h_style = ParagraphStyle('Head', parent=styles['Heading2'], fontSize=12, leading=15, textColor=colors.HexColor('#2563EB'), spaceBefore=10, spaceAfter=6)
        cell_bold = ParagraphStyle('CB', parent=styles['Normal'], fontSize=9, fontName='Helvetica-Bold', textColor=colors.HexColor('#0F172A'))
        cell_text = ParagraphStyle('CT', parent=styles['Normal'], fontSize=9, textColor=colors.HexColor('#334155'))

        story = [
            Paragraph(f"CodeSphere AI — {title}", title_style),
            Spacer(1, 10),
            Paragraph("Summary Metrics", h_style)
        ]

        metrics_data = [
            [Paragraph("Total Sessions", cell_bold), Paragraph(str(total_sessions), cell_text), Paragraph("Total Students", cell_bold), Paragraph(str(total_students), cell_text)],
            [Paragraph("Compiler Executions", cell_bold), Paragraph(str(total_runs), cell_text), Paragraph("Successful Executions", cell_bold), Paragraph(str(successful_runs), cell_text)],
            [Paragraph("Failed Executions", cell_bold), Paragraph(str(failed_runs), cell_text), Paragraph("Average Progress", cell_bold), Paragraph(f"{avg_progress}%", cell_text)],
            [Paragraph("Average AI Score", cell_bold), Paragraph(f"{avg_ai_score}/100", cell_text), Paragraph("Average Code Quality", cell_bold), Paragraph(f"{avg_code_quality}/100", cell_text)]
        ]
        t_metrics = Table(metrics_data, colWidths=[130, 130, 130, 130])
        t_metrics.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,-1), colors.HexColor('#F8FAFC')),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#CBD5E1')),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('PADDING', (0,0), (-1,-1), 6)
        ]))
        story.append(t_metrics)
        story.append(Spacer(1, 12))

        # Sessions Breakdown
        story.append(Paragraph("Sessions Overview", h_style))
        sess_rows = [[Paragraph("Date", cell_bold), Paragraph("Session Title", cell_bold), Paragraph("Language", cell_bold), Paragraph("Mode", cell_bold), Paragraph("Status", cell_bold)]]
        for s in sessions[:25]:
            d_str = s.created_at.strftime("%Y-%m-%d") if s.created_at else "N/A"
            sess_rows.append([
                Paragraph(d_str, cell_text),
                Paragraph(str(s.title), cell_text),
                Paragraph(str(s.language).upper(), cell_text),
                Paragraph(str(s.mode), cell_text),
                Paragraph(str(s.status).capitalize(), cell_text)
            ])
        if len(sess_rows) > 1:
            t_sess = Table(sess_rows, colWidths=[80, 200, 70, 90, 80])
            t_sess.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1E293B')),
                ('TEXTCOLOR', (0,0), (-1,0), colors.white),
                ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#E2E8F0')),
                ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
                ('PADDING', (0,0), (-1,-1), 5)
            ]))
            story.append(t_sess)

        doc.build(story)

    @classmethod
    def _build_excel_report(cls, filepath, title, total_sessions, total_students, total_runs, successful_runs, failed_runs, avg_progress, avg_ai_score, avg_code_quality, sessions, participations):
        wb = Workbook()
        ws = wb.active
        ws.title = "Summary"

        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        title_font = Font(name="Calibri", size=14, bold=True, color="1E293B")
        sub_font = Font(name="Calibri", size=11, bold=True, color="2563EB")

        ws.append([f"CodeSphere AI — {title}"])
        ws.cell(row=1, column=1).font = title_font
        ws.append([])

        ws.append(["AGGREGATE METRICS", "VALUE"])
        ws.cell(row=3, column=1).font = sub_font
        ws.append(["Total Sessions", total_sessions])
        ws.append(["Total Students", total_students])
        ws.append(["Total Code Executions", total_runs])
        ws.append(["Successful Executions", successful_runs])
        ws.append(["Failed Executions", failed_runs])
        ws.append(["Average Progress", f"{avg_progress}%"])
        ws.append(["Average AI Score", avg_ai_score])
        ws.append(["Average Code Quality", avg_code_quality])
        ws.append([])

        # Sessions Sheet
        ws2 = wb.create_sheet(title="Sessions")
        ws2.append(["Session ID", "Session Code", "Title", "Language", "Mode", "Status", "Created At"])
        for cell in ws2[1]:
            cell.font = header_font
            cell.fill = header_fill

        for s in sessions:
            ws2.append([
                s.id,
                s.session_code,
                s.title,
                s.language,
                s.mode,
                s.status,
                s.created_at.strftime("%Y-%m-%d %H:%M:%S UTC") if s.created_at else "N/A"
            ])

        for w in [ws, ws2]:
            for col in w.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = get_column_letter(col[0].column)
                w.column_dimensions[col_letter].width = max(max_len + 3, 12)

        wb.save(filepath)
